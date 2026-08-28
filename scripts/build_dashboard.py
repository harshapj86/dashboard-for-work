"""
Aptronix Service Dashboard — automated build script.

Reads the master Excel workbook (data/For_AI_Dashboard_.xlsx), runs the full
ETL pipeline (clean raw transactions, correct GP using the GP report,
aggregate into the fact tables the dashboard needs, and optionally fold in
Target sheets), then embeds the result into dashboard_template.html to
produce site/index.html — ready for GitHub Pages.

Run locally with:  python scripts/build_dashboard.py
The GitHub Actions workflow runs this automatically whenever
data/For_AI_Dashboard_.xlsx changes.
"""
import json
import os
import re
import pickle

import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'data', 'For_AI_Dashboard_.xlsx')
TEMPLATE = os.path.join(ROOT, 'scripts', 'dashboard_template.html')
APP_JS = os.path.join(ROOT, 'scripts', 'app.js')
OUT_DIR = os.path.join(ROOT, 'site')
OUT_FILE = os.path.join(OUT_DIR, 'index.html')

MONTH_NAMES = {m.lower(): i + 1 for i, m in enumerate(
    ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])}
MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def r2(x):
    return round(float(x), 2)


def parse_month(s):
    if pd.isna(s):
        return None
    parts = str(s).strip().split()
    if not parts or parts[0][:3].lower() not in MONTH_NAMES:
        return None
    mon = MONTH_NAMES[parts[0][:3].lower()]
    yr = int(parts[1])
    if yr < 100:
        yr += 2000
    return yr, mon


# ---------------------------------------------------------------
# 1. Load and clean raw transaction sheets
# ---------------------------------------------------------------
def load_clean():
    xl = pd.ExcelFile(SRC)
    dfs = []
    for sheet in xl.sheet_names:
        if not sheet.lower().replace(' ', '').startswith(('rawdata', 'rawdata24', 'rawdata2025', 'rawdata26')):
            continue
    # match sheets by fuzzy name (handles "Raw Data 24-25", "Rawdata 26-27" etc.)
    raw_sheets = [s for s in xl.sheet_names if 'raw' in s.lower() and 'data' in s.lower()]
    dfs = []
    for sheet in raw_sheets:
        d = pd.read_excel(xl, sheet_name=sheet)
        d.columns = [c.strip() for c in d.columns]
        if 'TXN Month' in d.columns:
            d = d.rename(columns={'TXN Month': 'TXNMonth'})
        dfs.append(d)
    df = pd.concat(dfs, ignore_index=True, sort=False)
    df = df[df['Status'] != 'Cancelled'].copy()

    df = df[df['TXNMonth'].apply(lambda s: parse_month(s) is not None)].copy()
    parsed = df['TXNMonth'].apply(parse_month)
    df['Yr'] = parsed.apply(lambda x: x[0])
    df['Mon'] = parsed.apply(lambda x: x[1])
    df['MonthKey'] = df['Yr'] * 100 + df['Mon']
    df['MonthLabel'] = df.apply(lambda r: f"{MONTH_ABBR[r['Mon']-1]} {str(r['Yr'])[2:]}", axis=1)
    df['FY'] = df.apply(lambda r: f"FY{str(r['Yr'])[2:]}-{str(r['Yr']+1)[2:]}" if r['Mon'] >= 4
                         else f"FY{str(r['Yr']-1)[2:]}-{str(r['Yr'])[2:]}", axis=1)
    df['Quarter'] = df['Mon'].apply(lambda m: 'Q1' if m in (4, 5, 6) else 'Q2' if m in (7, 8, 9)
                                     else 'Q3' if m in (10, 11, 12) else 'Q4')
    df['FYQ'] = df['FY'] + ' ' + df['Quarter']

    loc = pd.read_excel(xl, sheet_name='Location Master')
    loc.columns = [c.strip() for c in loc.columns]
    loc_map = loc.set_index('Centre Name').to_dict('index')

    def get_loc(branch, field, default):
        if branch in loc_map:
            v = loc_map[branch].get(field)
            return v if pd.notna(v) else default
        return default

    unmapped_type = lambda b: 'Repair Drop Off' if 'drop off' in b.lower() else 'Service Centre'

    df['State'] = df['Branch ID'].apply(lambda b: get_loc(b, 'State', 'Unmapped'))
    df['ARM'] = df['Branch ID'].apply(lambda b: get_loc(b, 'ARM', 'Unmapped'))
    df['LocationType'] = df.apply(lambda r: get_loc(r['Branch ID'], 'Location Type', unmapped_type(r['Branch ID'])), axis=1)
    df['Centre'] = df['Branch ID']

    df['Revenue'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['GP'] = pd.to_numeric(df['GP AT Amount'], errors='coerce').fillna(0)

    BU_NORMALIZE = {'incident fees': 'Incident Fees', 'accessory repairs': 'Accessory Repairs'}
    df['BU'] = df['Business Unit'].fillna('Other').apply(lambda x: BU_NORMALIZE.get(str(x).strip().lower(), x))
    df['Category'] = df['Category'].fillna('Other')
    df['ProductFamily'] = df['Product Family'].fillna('Other')
    df['Item_Name'] = df['Item_Name'].fillna('Unspecified Item')
    df['TxnType'] = df['Transaction Type'].fillna('Other')
    df['Executive'] = df['Executive'].fillna('Unassigned')

    return df, loc, xl


# ---------------------------------------------------------------
# 2. GP correction from the GP report sheet (2 stacked FY blocks)
# ---------------------------------------------------------------
def extract_gp_true():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if 'GP ' not in wb.sheetnames:
        return {}
    ws = wb['GP ']

    def extract_block(header_row, data_start, data_end):
        header = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        month_cols = header[1:13]
        month_keys = []
        for m in month_cols:
            pm = parse_month(m)
            month_keys.append(pm[0]*100+pm[1] if pm else None)
        out = {}
        for row in ws.iter_rows(min_row=data_start, max_row=data_end, values_only=True):
            branch = row[0]
            if branch is None or 'grand total' in str(branch).lower() or 'total gp' in str(branch).lower():
                continue
            for i, mk in enumerate(month_keys):
                if mk is None:
                    continue
                val = row[1 + i]
                if val is None:
                    continue
                out[(branch, mk)] = float(val)
        return out

    gp_true = {}
    # First block starts at row 3 (header), data rows follow until a blank/Grand Total row.
    # We scan for header rows containing 12 month-like labels rather than hardcoding row numbers,
    # so this survives minor sheet edits.
    header_rows = []
    for r in range(1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if row[0] == 'Branch ID' and row[1] is not None:
            header_rows.append(r)
    for hr in header_rows:
        # find extent of this block: rows until next header row or sheet end
        next_header = min([h for h in header_rows if h > hr], default=ws.max_row + 1)
        data_end = next_header - 2  # leave room for a trailing Grand Total row
        gp_true.update(extract_block(hr, hr + 1, data_end))
    return gp_true


# ---------------------------------------------------------------
# 3. Targets (optional sheets: "Revenue Target", "GP Target", "CSAT Target")
# ---------------------------------------------------------------
def extract_targets(xl):
    targets = {'revenue': {}, 'gp': {}, 'csat': {}}
    sheet_map = {'Revenue Target': 'revenue', 'GP Target': 'gp', 'CSAT Target': 'csat'}
    for sheet_name, key in sheet_map.items():
        if sheet_name not in xl.sheet_names:
            continue
        d = pd.read_excel(xl, sheet_name=sheet_name, header=3)  # header row 4 (1-indexed) in our template
        d.columns = [str(c).strip() for c in d.columns]
        if 'Branch ID' not in d.columns:
            continue
        month_cols = [c for c in d.columns if c != 'Branch ID']
        for _, row in d.iterrows():
            branch = row['Branch ID']
            if pd.isna(branch):
                continue
            for mc in month_cols:
                val = row[mc]
                if pd.isna(val):
                    continue
                pm = parse_month(mc)
                if pm is None:
                    continue
                yr, mon = pm
                mk = yr * 100 + mon
                targets[key][(branch, mk)] = r2(float(val))
    return targets


# ---------------------------------------------------------------
# 4. Aggregate into the fact tables the dashboard expects
# ---------------------------------------------------------------
def build_payload():
    df, loc, xl = load_clean()
    gp_true = extract_gp_true()
    targets = extract_targets(xl)

    raw_gp_by_cm = df.groupby(['Centre', 'MonthKey'])['GP'].sum().to_dict()
    correction = {}
    for (centre, mk), raw_sum in raw_gp_by_cm.items():
        true_val = gp_true.get((centre, mk))
        if true_val is not None and raw_sum != 0:
            correction[(centre, mk)] = true_val / raw_sum
        else:
            correction[(centre, mk)] = 1.0
    df['GPFactor'] = df.apply(lambda r: correction.get((r['Centre'], r['MonthKey']), 1.0), axis=1)
    df['GPCorrected'] = df['GP'] * df['GPFactor']
    GPCOL = 'GPCorrected'

    gp_covered_months = sorted(set(int(mk) for (_, mk) in gp_true.keys()))

    months = df[['MonthKey', 'MonthLabel', 'FY', 'Quarter', 'FYQ']].drop_duplicates().sort_values('MonthKey')
    month_list = months.to_dict('records')
    for m in month_list:
        m['MonthKey'] = int(m['MonthKey'])

    centre_attrs = df.groupby('Centre').agg(State=('State', 'first'), ARM=('ARM', 'first'),
                                             LocationType=('LocationType', 'first'),
                                             City=('Branch City', 'first')).reset_index()
    centres = centre_attrs.to_dict('records')

    def agg(group_cols):
        g = df.groupby(group_cols, as_index=False).agg(Revenue=('Revenue', 'sum'), GP=(GPCOL, 'sum'), Txns=('Revenue', 'count'))
        return g

    g = agg(['MonthKey', 'Centre'])
    fact_month = [[int(r['MonthKey']), r['Centre'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = agg(['MonthKey', 'Centre', 'BU'])
    fact_bu = [[int(r['MonthKey']), r['Centre'], r['BU'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = agg(['MonthKey', 'Centre', 'Category'])
    fact_cat = [[int(r['MonthKey']), r['Centre'], r['Category'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = agg(['MonthKey', 'Centre', 'ProductFamily'])
    fact_pf = [[int(r['MonthKey']), r['Centre'], r['ProductFamily'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = agg(['MonthKey', 'Centre', 'Item_Name'])
    fact_item = [[int(r['MonthKey']), r['Centre'], r['Item_Name'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = agg(['MonthKey', 'Centre', 'TxnType'])
    fact_txntype = [[int(r['MonthKey']), r['Centre'], r['TxnType'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    g = df.groupby(['Centre', 'Executive'], as_index=False).agg(Revenue=('Revenue', 'sum'), GP=(GPCOL, 'sum'), Txns=('Revenue', 'count'))
    fact_exec = [[r['Centre'], r['Executive'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    df['DateStr'] = pd.to_datetime(df['TXNDate'], format='%d-%m-%Y').dt.strftime('%Y-%m-%d')
    g = df.groupby(['DateStr', 'Centre'], as_index=False).agg(Revenue=('Revenue', 'sum'), GP=(GPCOL, 'sum'), Txns=('Revenue', 'count'))
    fact_day = [[r['DateStr'], r['Centre'], r2(r['Revenue']), r2(r['GP']), int(r['Txns'])] for _, r in g.iterrows()]

    csat_rows = []
    if 'CSAT' in xl.sheet_names:
        csat_raw = pd.read_excel(xl, sheet_name='CSAT')
        csat_raw.columns = [str(c).strip() for c in csat_raw.columns]
        date_cols = [c for c in csat_raw.columns if c not in ('Ship-To', 'Metrics')]
        for _, row in csat_raw.iterrows():
            cname = str(row['Ship-To']).strip()
            if cname == 'Average':
                continue
            for dc in date_cols:
                val = row[dc]
                if pd.isna(val):
                    continue
                try:
                    fval = float(val)
                except Exception:
                    continue
                try:
                    dt = pd.to_datetime(dc)
                except Exception:
                    continue
                mk = dt.year * 100 + dt.month
                csat_rows.append([mk, cname, r2(fval * 100 if fval <= 1 else fval)])

    # Targets, flattened to [mk, centre, value] lists (same shape as csat) for easy JS consumption
    def flatten_targets(d):
        return [[mk, centre, val] for (centre, mk), val in d.items()]

    payload = {
        'months': month_list,
        'gpCoveredMonths': gp_covered_months,
        'centres': centres,
        'fact_month': fact_month,
        'fact_bu': fact_bu,
        'fact_cat': fact_cat,
        'fact_pf': fact_pf,
        'fact_item': fact_item,
        'fact_txntype': fact_txntype,
        'fact_exec': fact_exec,
        'fact_day': fact_day,
        'csat': csat_rows,
        'target_revenue': flatten_targets(targets['revenue']),
        'target_gp': flatten_targets(targets['gp']),
        'target_csat': flatten_targets(targets['csat']),
    }
    return payload


def main():
    payload = build_payload()
    data_json = json.dumps(payload, separators=(',', ':'))

    with open(TEMPLATE) as f:
        html = f.read()
    with open(APP_JS) as f:
        js = f.read()

    html = re.sub(
        r'(<script id="data-blob" type="application/json">).*?(</script>)',
        lambda m: m.group(1) + data_json + m.group(2),
        html, flags=re.S,
    )
    html = html.replace('__APP_JS__', js)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT_FILE, 'w') as f:
        f.write(html)
    print(f"Wrote {OUT_FILE} ({len(html)/1e6:.2f} MB)")


if __name__ == '__main__':
    main()
